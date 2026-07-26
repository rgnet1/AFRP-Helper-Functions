import os
import sys
import gc
import warnings
import re
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Type, Set

import pandas as pd
import pytz
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils.badges.data_store import BadgeDataStore
from utils.badges.event_statistics import EventStatisticsReport
from utils.badges.meal_options import (
    MEAL_PREFERENCE_COLUMN,
    MEAL_QUESTION_EXCLUDE,
    MEAL_QUESTION_INCLUDE,
    build_meal_preference_value,
    default_source_config,
)
from utils.badges.event_preprocessing.default import DefaultPreprocessing
from utils.badges.pre_processing_module import PreprocessingConfig, PreprocessingBase
from utils.badges.file_validator import FileValidator, FileTypes


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s: %(message)s'
)
logger = logging.getLogger(__name__)

# Suppress openpyxl UserWarning about default style
warnings.simplefilter("ignore", UserWarning)

class RegistrationColumns:
    CONTACT_ID = "Contact ID (Existing Contact) (Contact)"
    MEMBER_ID = "Member ID (Existing Contact) (Contact)"
    FIRST_NAME = "First Name (Existing Contact) (Contact)"
    MIDDLE_NAME = "Middle Name (Existing Contact) (Contact)"
    LAST_NAME = "Last Name (Existing Contact) (Contact)"
    TITLE = "Title (Existing Contact) (Contact)"
    MAIDEN_NAME = "Maiden Name (Existing Contact) (Contact)"
    LOCAL_CLUB = "Local Club (Existing Contact) (Contact)"
    GENDER = "Gender (Existing Contact) (Contact)"
    AGE = "Age (Existing Contact) (Contact)"
    EVENT = "Event"
    STATUS = "Status Reason"
    CREATED_ON = "Created On"
    HOUSEHOLD_ID = "Household ID (Existing Contact) (Contact)"
    HOUSEHOLD = "Household (Existing Contact) (Contact)"
    HEAD_OF_HOUSEHOLD = "Head of Household (Existing Contact) (Contact)"
    
    # Map of standardized names to possible column names in file
    MAPPINGS = {
        'Contact ID': [CONTACT_ID, 'Contact ID', 'Contact'],
        'Member ID': [MEMBER_ID, 'Member ID', 'ID'],
        'First Name': [FIRST_NAME],
        'Middle Name': [MIDDLE_NAME, 'Middle Name'],
        'Last Name': [LAST_NAME],
        'Title': [TITLE],
        'Maiden Name': [MAIDEN_NAME, 'Maiden Name'],
        'Local Club': [LOCAL_CLUB],
        'Gender': [GENDER],
        'Age': [AGE],
        'Household ID': [HOUSEHOLD_ID, 'Household ID'],
        'Household': [HOUSEHOLD, 'Household'],
        'Head of Household': [HEAD_OF_HOUSEHOLD, 'Head of Household'],
        'Event': [EVENT, 'Event '],  # Note the space variant
        'Status': [STATUS],
        'Created On': [CREATED_ON, 'Created On', 'CreatedOn', 'Date Created']
    }

class SeatingColumns:
    CONTACT_ID = "Contact ID (Contact) (Contact)"
    EVENT = "Event"  # Use the Event column from seating chart
    TABLE = "Table"
    
    # Map of standardized names to possible column names in file
    MAPPINGS = {
        'Contact ID': [CONTACT_ID, 'Contact ID', 'Contact'],
        'Event': [EVENT],  # Only use Event column
        'Table': [TABLE]
    }

class QRCodeColumns:
    CONTACT_ID = "Contact ID (Event Guest Contact Id) (Contact)"
    QR_CODE = "QR Code Value"
    
    # Map of standardized names to possible column names in file
    MAPPINGS = {
        'Contact ID': [CONTACT_ID, 'Contact ID', 'Contact', 'Event Guest Contact Id'],
        'QR Code': [QR_CODE, 'QR Code', 'QR Code Value']
    }

class FormResponseColumns:
    CONTACT_ID = "Contact ID (Contact) (Contact)"
    EVENT = "Campaign"  # Use Campaign column from form responses
    QUESTION = "Form Question"
    RESPONSE = "Guest Response"
    
    # Map of standardized names to possible column names in file
    MAPPINGS = {
        'Contact ID': [CONTACT_ID, 'Contact ID', 'Contact'],
        'Event': [EVENT],  # Only use Campaign column
        'Question': [QUESTION, 'Form Question', 'Question'],
        'Response': [RESPONSE, 'Guest Response', 'Response', 'Answer']
    }

class EventRegistrationProcessorV3:
    def __init__(self, config: Optional[PreprocessingConfig] = None, preprocessor_class: Optional[Type[PreprocessingBase]] = None):
        """
        Initialize the processor with optional configuration and preprocessor class.
        
        Args:
            config: Optional configuration for preprocessing
            preprocessor_class: Optional class to use for preprocessing. If not provided, defaults to DefaultPreprocessing
        """
        self.config = config
        if preprocessor_class is None:
            logger.debug("No preprocessor class provided, defaulting to DefaultPreprocessing")
            preprocessor_class = DefaultPreprocessing
        
        logger.debug(f"Initializing preprocessor with class: {preprocessor_class.__name__}")
        self.preprocessor = preprocessor_class(config)
        self.stats_reporter = EventStatisticsReport()
        self._registration_created_on: Optional[pd.DataFrame] = None
        
    def find_latest_files(self) -> Dict[str, str]:
        """Find the latest version of each file type in the directory."""
        return FileValidator.find_latest_files()

    def _find_column_name(self, df: pd.DataFrame, column_mappings: Dict[str, List[str]], required_column: str) -> Optional[str]:
        """Find the actual column name in the DataFrame based on possible mappings."""
        possible_names = column_mappings[required_column]
        for name in possible_names:
            if name in df.columns:
                return name
        return None

    def _standardize_columns(self, df: pd.DataFrame, column_mappings: Dict[str, List[str]]) -> Tuple[pd.DataFrame, List[str]]:
        """Standardize column names based on mappings and return missing columns."""
        # Clean column names - ensure all are strings first
        df.columns = df.columns.astype(str).str.strip()
        
        # Create reverse mapping
        column_mapping = {}
        missing_columns = []
        
        for standard_name, possible_names in column_mappings.items():
            found_name = self._find_column_name(df, column_mappings, standard_name)
            if found_name:
                column_mapping[found_name] = standard_name
            else:
                missing_columns.append(standard_name)
        
        # Rename columns to standard names
        if column_mapping:
            df = df.rename(columns=column_mapping)
            
        return df, missing_columns

    def process_registration_data(self, reg_df: pd.DataFrame) -> pd.DataFrame:
        """Process the main registration data."""
        logger.debug("Registration file has %d columns", len(reg_df.columns))
        
        # Standardize column names
        reg_df, missing_columns = self._standardize_columns(reg_df, RegistrationColumns.MAPPINGS)
        if 'Created On' in reg_df.columns:
            self._registration_created_on = (
                reg_df[['Contact ID', 'Created On']]
                .drop_duplicates(subset=['Contact ID'])
                .copy()
            )
        # Middle/Maiden Name are optional - older data files may not include them
        optional_columns = {'Middle Name', 'Maiden Name', 'Household ID', 'Household', 'Head of Household'}
        for col in [c for c in missing_columns if c in optional_columns]:
            logger.info(f"Optional column '{col}' not found in registration data, adding empty column")
            reg_df[col] = ''
        missing_columns = [c for c in missing_columns if c not in optional_columns]
        if missing_columns:
            logger.debug("Available registration columns: %s", list(reg_df.columns))
            raise ValueError(f"Missing required columns in registration data: {', '.join(missing_columns)}")
        
        # Filter for paid registrations
        paid_df = reg_df[reg_df['Status'] == 'Paid']
        logger.info(f"\nFound {len(paid_df)} paid registrations out of {len(reg_df)} total")
        
        # Get unique events - handle both 'Event' and 'Event ' column names
        event_col = 'Event' if 'Event' in paid_df.columns else 'Event '
        unique_events = [e for e in paid_df[event_col].unique() if pd.notna(e)]
        logger.info("Found %d unique events", len(unique_events))
        logger.debug("Events: %s", unique_events)
        
        # Create base DataFrame with unique contacts using only the key identifying columns
        unique_columns = [
            'Contact ID', 'Member ID', 'First Name', 'Middle Name', 'Last Name', 'Maiden Name',
            'Title', 'Local Club', 'Gender', 'Age',
            'Household ID', 'Household', 'Head of Household',
        ]
        present_columns = [col for col in unique_columns if col in paid_df.columns]
        transformed_df = paid_df[present_columns].drop_duplicates(subset=['Contact ID']).reset_index(drop=True)
        
        logger.info(f"\nFound {len(transformed_df)} unique contacts")

        # Format names to proper case
        logger.info("Formatting names to proper case...")
        transformed_df['First Name'] = transformed_df['First Name'].apply(lambda x: str(x).strip().title() if pd.notna(x) else x)
        transformed_df['Middle Name'] = transformed_df['Middle Name'].apply(lambda x: str(x).strip().title() if pd.notna(x) and str(x).strip() else '')
        transformed_df['Last Name'] = transformed_df['Last Name'].apply(lambda x: str(x).strip().title() if pd.notna(x) else x)
        transformed_df['Maiden Name'] = transformed_df['Maiden Name'].apply(lambda x: str(x).strip().title() if pd.notna(x) and str(x).strip() else '')
        
        # Normalize Gender values - handle cases where formatted values didn't come through
        logger.info("Normalizing Gender values...")
        def normalize_gender(value):
            if pd.isna(value) or value == '' or str(value).strip() == '':
                # Blank/null in CRM typically means Female (option value 2 with no label)
                return 'Female'
            value_str = str(value).strip()
            # Handle numeric codes from CRM option sets
            if value_str == '1':
                return 'Male'
            elif value_str == '2':
                return 'Female'
            # Handle text values (already formatted or from old data)
            elif value_str.lower() in ['male', 'm']:
                return 'Male'
            elif value_str.lower() in ['female', 'f']:
                return 'Female'
            # If already properly formatted, keep it
            elif value_str in ['Male', 'Female']:
                return value_str
            # Log unknown values but default to blank to avoid incorrect data
            else:
                logger.warning(f"Unknown gender value: '{value}', leaving blank")
                return ''
        
        transformed_df['Gender'] = transformed_df['Gender'].apply(normalize_gender)
        gender_counts = transformed_df['Gender'].value_counts().to_dict()
        logger.info(f"Gender distribution after normalization: {gender_counts}")
        
        # Vectorized event registration columns (contact x event -> event name if registered)
        reg_pairs = paid_df[['Contact ID', event_col]].drop_duplicates()
        reg_pairs = reg_pairs.rename(columns={event_col: '_event_name'})
        reg_pairs['_registered'] = reg_pairs['_event_name']
        event_wide = reg_pairs.pivot(
            index='Contact ID', columns='_event_name', values='_registered'
        )
        for event in unique_events:
            if event in event_wide.columns:
                transformed_df[event] = transformed_df['Contact ID'].map(event_wide[event])
            else:
                transformed_df[event] = None
        
        return transformed_df

    def add_seating_info(
        self,
        df: pd.DataFrame,
        seating_df: pd.DataFrame,
        contact_ids: Optional[Set] = None,
    ) -> pd.DataFrame:
        """Add seating information for each event."""
        if seating_df.empty or len(seating_df) == 0:
            logger.info("No seating data found - skipping table assignment columns")
            return df
        
        logger.debug("Seating file columns: %s", list(seating_df.columns))
        
        seating_df.columns = seating_df.columns.astype(str).str.strip()
        seating_df, missing_columns = self._standardize_columns(seating_df, SeatingColumns.MAPPINGS)
        if missing_columns:
            logger.warning("Missing required columns in seating data: %s", missing_columns)
            return df

        if contact_ids is not None:
            seating_df = seating_df[seating_df['Contact ID'].isin(contact_ids)]
            if seating_df.empty:
                return df
        
        logger.info("Processing seating assignments...")
        if 'Created On' in seating_df.columns:
            seating_info = (
                seating_df.sort_values('Created On', ascending=False)
                .drop_duplicates(['Contact ID', 'Event'])
            )
        else:
            seating_info = seating_df.drop_duplicates(['Contact ID', 'Event'])
        
        events_with_seating = sorted(
            e for e in seating_df['Event'].unique() if pd.notna(e) and str(e).strip() != ''
        )
        for event in events_with_seating:
            df[f"{event} ~ Table"] = ''
        
        seating_info = seating_info.copy()
        seating_info['_table'] = seating_info['Table'].astype(str).str.strip()
        seating_info = seating_info[
            seating_info['_table'].notna()
            & seating_info['_table'].ne('')
            & seating_info['_table'].ne('nan')
        ]
        if seating_info.empty:
            return df

        seating_info['column_name'] = seating_info['Event'].astype(str) + ' ~ Table'
        table_pivot = seating_info.pivot_table(
            index='Contact ID', columns='column_name', values='_table', aggfunc='first'
        )

        for col in table_pivot.columns:
            event_name = str(col).replace(' ~ Table', '')
            if event_name not in df.columns:
                continue
            registered = df[event_name].notna() & (df[event_name].astype(str).str.strip() != '')
            if not registered.any():
                continue
            mapped = df.loc[registered, 'Contact ID'].map(table_pivot[col])
            df.loc[registered, col] = mapped.fillna('').astype(str)

        logger.info("Table assignment summary:")
        for event in events_with_seating:
            column_name = f"{event} ~ Table"
            assigned = (df[column_name].astype(str).str.strip() != '').sum()
            logger.info("  - %s: %d assignments", event, assigned)
        
        return df

    def add_form_responses(
        self,
        df: pd.DataFrame,
        forms_df: pd.DataFrame,
        contact_ids: Optional[Set] = None,
    ) -> pd.DataFrame:
        """Add form responses for each event."""
        if forms_df.empty or len(forms_df) == 0:
            logger.info("No form responses data found - skipping form response columns")
            return df
        
        logger.debug("Form responses columns: %d", len(forms_df.columns))
        
        forms_df, missing_columns = self._standardize_columns(forms_df, FormResponseColumns.MAPPINGS)
        if missing_columns:
            logger.warning("Missing required form response columns: %s", missing_columns)
            return df

        if contact_ids is not None:
            forms_df = forms_df[forms_df['Contact ID'].isin(contact_ids)]
            if forms_df.empty:
                return df
        
        # Ensure Created On is properly parsed as datetime
        try:
            forms_df['Created On'] = pd.to_datetime(forms_df['Created On'])
        except Exception as e:
            logger.warning(f"Could not parse Created On column: {str(e)}")
            # If we can't parse Created On, we'll just use the first response for each contact
            forms_df['Created On'] = pd.Timestamp.now()
        
        event_questions = forms_df.groupby('Event')['Question'].unique()
        logger.info("Found form questions for %d events", len(event_questions))
        
        for event in event_questions.index:
            for question in event_questions[event]:
                column_name = f"{event} ~ {question}"
                
                event_question_responses = forms_df[
                    (forms_df['Event'] == event) &
                    (forms_df['Question'] == question)
                ]
                
                duplicates = event_question_responses.groupby('Contact ID').size()
                dup_count = int((duplicates > 1).sum())
                if dup_count:
                    logger.warning(
                        "Found duplicate responses for %s - %s (%d contacts)",
                        event,
                        question,
                        dup_count,
                    )
                
                latest_responses = (
                    event_question_responses
                    .sort_values('Created On', ascending=False)
                    .groupby('Contact ID', as_index=False)
                    .first()
                )
                response_dict = dict(
                    zip(latest_responses['Contact ID'], latest_responses['Response'])
                )
                df[column_name] = df['Contact ID'].map(response_dict)
        
        return df

    def _find_meal_preference_columns(self, df: pd.DataFrame) -> List[str]:
        """Return per-event form-response columns that represent a meal choice.

        Columns look like ``{event} ~ {question}``; we keep the ones whose
        question matches a meal-preference keyword and skip logistics questions
        (children's meal, allergies, t-shirt size, ...).
        """
        meal_cols = []
        for col in df.columns:
            if " ~ " not in str(col):
                continue
            question = str(col).split(" ~ ", 1)[1]
            if MEAL_QUESTION_INCLUDE.search(question) and not MEAL_QUESTION_EXCLUDE.search(question):
                meal_cols.append(col)
        return meal_cols

    def merge_meal_preferences(self, df: pd.DataFrame) -> pd.DataFrame:
        """Build a default merged ``Meal Preference`` column for Excel export.

        Uses registered-banquet-only logic (one banquet room per attendee).
        Non-banquet meal questions stay as separate ``{event} ~ {question}``
        columns; preprocessing templates can enable them via meal_preference_sources.
        """
        meal_cols = self._find_meal_preference_columns(df)
        if not meal_cols:
            logger.info("No meal-preference columns found - skipping meal merge")
            return df

        questions = []
        for index, col in enumerate(meal_cols):
            event_name, question = str(col).split(" ~ ", 1)
            questions.append(
                {
                    "campaign_name": event_name,
                    "question": question,
                    "order": index,
                }
            )
        mappings = None
        sources_config = None
        if self.config:
            mappings = getattr(self.config, 'meal_preference_mappings', None) or None
            sources_config = getattr(self.config, 'meal_preference_sources', None) or None
        if not sources_config:
            sources_config = default_source_config(questions)

        logger.info("Merging meal-preference columns into '%s':", MEAL_PREFERENCE_COLUMN)
        for col in meal_cols:
            logger.info("  - %s", col)

        merged = df.apply(
            lambda row: build_meal_preference_value(
                row.to_dict(), sources_config, mappings
            ),
            axis=1,
        )
        if MEAL_PREFERENCE_COLUMN in df.columns:
            existing = df[MEAL_PREFERENCE_COLUMN].astype(str).str.strip()
            df[MEAL_PREFERENCE_COLUMN] = existing.where(existing != '', merged)
        else:
            df[MEAL_PREFERENCE_COLUMN] = merged

        assigned = df[df[MEAL_PREFERENCE_COLUMN].astype(str).str.strip() != ''].shape[0]
        logger.info(
            "Meal preference assigned for %d of %d contacts", assigned, len(df)
        )
        return df

    def add_qr_codes(
        self,
        df: pd.DataFrame,
        qr_df: pd.DataFrame,
        contact_ids: Optional[Set] = None,
    ) -> pd.DataFrame:
        """Add QR code information."""
        if qr_df.empty or len(qr_df) == 0:
            logger.info("No QR code data found - skipping QR code column")
            df['QR Code'] = ''
            return df
        
        logger.debug("QR codes file columns: %d", len(qr_df.columns))
        
        qr_df, missing_columns = self._standardize_columns(qr_df, QRCodeColumns.MAPPINGS)
        if missing_columns:
            logger.warning("Missing required QR code columns: %s", missing_columns)
            df['QR Code'] = ''
            return df

        if contact_ids is not None:
            qr_df = qr_df[qr_df['Contact ID'].isin(contact_ids)]
            if qr_df.empty:
                df['QR Code'] = ''
                return df
        
        # Ensure Created On is properly parsed as datetime
        try:
            qr_df['Created On'] = pd.to_datetime(qr_df['(Do Not Modify) Modified On'])
        except Exception as e:
            logger.warning(f"Could not parse Created On column: {str(e)}")
            # If we can't parse Created On, we'll just use the first QR code for each contact
            qr_df['Created On'] = pd.Timestamp.now()
        
        duplicates = qr_df.groupby('Contact ID').size()
        dup_count = int((duplicates > 1).sum())
        if dup_count:
            logger.warning("Found duplicate QR codes for %d contacts", dup_count)
        # Keep only the most recent QR code for each contact
        latest_qr_codes = (qr_df
                          .sort_values('Created On', ascending=False)
                          .groupby('Contact ID', as_index=False)
                          .first())
        
        # Create a dictionary mapping Contact ID to QR Code instead of using Series
        qr_code_dict = dict(zip(latest_qr_codes['Contact ID'], latest_qr_codes['QR Code']))
        
        # Map QR codes to Contact IDs using the dictionary
        df['QR Code'] = df['Contact ID'].map(qr_code_dict)
        logger.info(f"Added QR codes for {len(qr_code_dict)} contacts")
        return df

    def _enrich_households_if_needed(self, df: pd.DataFrame) -> pd.DataFrame:
        """Load household fields from file cache (CRM fetch on cache miss only)."""
        if not self.config or not getattr(self.config, 'group_by_household', False):
            return df
        cache_path = getattr(self.config, 'household_cache_path', None)
        if not cache_path:
            logger.warning("group_by_household enabled but household_cache_path not configured")
            return df
        from utils.badges.household_cache import enrich_dataframe
        from utils.dynamics_crm import DynamicsCRMClient
        try:
            crm_client = DynamicsCRMClient()
        except Exception as exc:
            logger.warning("Could not initialize CRM client for household cache: %s", exc)
            crm_client = None
        return enrich_dataframe(df, crm_client, cache_path)

    def _apply_attendee_sort(self, df: pd.DataFrame) -> pd.DataFrame:
        """Sort attendees for badge print order."""
        if df.empty:
            return df
        group_by_household = (
            self.config is not None
            and getattr(self.config, 'group_by_household', False)
        )
        if not group_by_household:
            return df.sort_values(by=['Last Name', 'First Name']).reset_index(drop=True)

        for col in ('Household ID', 'Household', 'Head of Household'):
            if col not in df.columns:
                df[col] = ''

        has_household_data = (
            df['Household ID'].astype(str).str.strip().ne('').any()
            or df['Household'].astype(str).str.strip().ne('').any()
        )
        if not has_household_data:
            logger.warning(
                "Group by household enabled but no household data available; using Last/First sort"
            )
            return df.sort_values(by=['Last Name', 'First Name']).reset_index(drop=True)

        work = df.copy()
        work['_hh_key'] = work['Household ID'].replace('', pd.NA).fillna(work['Contact ID'])
        work['_hh_sort'] = work['Household'].replace('', pd.NA).fillna(work['Last Name'])
        work['_head_sort'] = work['Head of Household'].map({'Yes': 0, 'No': 1}).fillna(1)
        work = work.sort_values(
            by=['_hh_sort', '_hh_key', '_head_sort', 'Last Name', 'First Name']
        )
        return work.drop(columns=['_hh_key', '_hh_sort', '_head_sort']).reset_index(drop=True)

    def _load_source_frames(self, directory: str = '.') -> Dict[str, pd.DataFrame]:
        """Load CRM source files (Parquet preferred) with memory-aware strategy."""
        paths = BadgeDataStore.find_source_paths(directory)
        frames = BadgeDataStore.load_all(paths)
        return {
            FileTypes.REGISTRATION: frames[FileTypes.REGISTRATION],
            FileTypes.SEATING: frames[FileTypes.SEATING],
            FileTypes.QR_CODES: frames[FileTypes.QR_CODES],
            FileTypes.FORM_RESPONSES: frames[FileTypes.FORM_RESPONSES],
        }

    def _sub_event_name(self) -> Optional[str]:
        if self.config and getattr(self.config, 'sub_event', None):
            return self.config.sub_event
        return None

    def _early_sub_event_contact_ids(self, result_df: pd.DataFrame) -> Optional[Set]:
        sub_event = self._sub_event_name()
        if not sub_event or sub_event not in result_df.columns:
            return None
        ids = result_df.loc[result_df[sub_event].notna(), 'Contact ID'].unique()
        return set(ids) if len(ids) else set()

    def _apply_sub_event_filters(self, result_df: pd.DataFrame) -> pd.DataFrame:
        sub_event = self._sub_event_name()
        if not sub_event:
            return result_df

        logger.info("Filtering data for sub-event: %s", sub_event)
        if sub_event not in result_df.columns:
            logger.warning("Sub-event column '%s' not found in DataFrame", sub_event)
            return pd.DataFrame(columns=result_df.columns)

        sub_event_contacts = result_df[result_df[sub_event].notna()]['Contact ID'].unique()
        if len(sub_event_contacts) == 0:
            logger.warning("No contacts found for sub-event: %s", sub_event)
            return pd.DataFrame(columns=result_df.columns)

        result_df = result_df[result_df['Contact ID'].isin(sub_event_contacts)].copy()
        logger.info("Found %d contacts registered for %s", len(result_df), sub_event)

        contact_columns = [
            'Contact ID', 'First Name', 'Middle Name', 'Last Name', 'Maiden Name',
            'Title', 'Local Club', 'Gender', 'Age', MEAL_PREFERENCE_COLUMN,
        ]
        relevant_columns = [col for col in contact_columns if col in result_df.columns]
        if sub_event in result_df.columns:
            relevant_columns.append(sub_event)
        for col in result_df.columns:
            if col.startswith(f"{sub_event} ~"):
                relevant_columns.append(col)

        logger.info(
            "Filtered to %d relevant columns for %s",
            len(relevant_columns),
            sub_event,
        )
        return result_df[relevant_columns]

    def _date_filter_contact_ids(self) -> Optional[Set]:
        if not self.config or not getattr(self.config, 'created_on_datetime', None):
            return None
        if self._registration_created_on is None or self._registration_created_on.empty:
            logger.warning("Created On not available in registration data - skipping date filter")
            return None
        reg_df = self._registration_created_on.copy()
        try:
            reg_df['Created On'] = pd.to_datetime(reg_df['Created On'])
            if reg_df['Created On'].dt.tz is None:
                reg_df['Created On'] = reg_df['Created On'].dt.tz_localize(self.config.tz)
            else:
                reg_df['Created On'] = reg_df['Created On'].dt.tz_convert(self.config.tz)
            filtered = reg_df[reg_df['Created On'] >= self.config.created_on_datetime]
            ids = set(filtered['Contact ID'].unique())
            logger.info(
                "Date filter matched %d contacts on or after %s",
                len(ids),
                self.config.created_on_datetime,
            )
            return ids
        except Exception as exc:
            logger.warning("Could not apply date filter: %s", exc)
            return None

    def transform_and_merge(self, source_directory: str = '.') -> pd.DataFrame:
        """Main function to transform and merge all data sources."""
        try:
            sources = self._load_source_frames(source_directory)
            reg_df = sources[FileTypes.REGISTRATION]
            seating_df = sources[FileTypes.SEATING]
            qr_df = sources[FileTypes.QR_CODES]
            forms_df = sources[FileTypes.FORM_RESPONSES]

            result_df = self.process_registration_data(reg_df)
            del reg_df
            gc.collect()

            contact_ids = self._early_sub_event_contact_ids(result_df)
            if contact_ids is not None:
                if not contact_ids:
                    return pd.DataFrame(columns=result_df.columns)
                result_df = result_df[result_df['Contact ID'].isin(contact_ids)].copy()
                logger.info(
                    "Early sub-event filter: %d contacts before companion merges",
                    len(result_df),
                )

            result_df = self.add_seating_info(result_df, seating_df, contact_ids)
            del seating_df
            gc.collect()

            result_df = self.add_form_responses(result_df, forms_df, contact_ids)
            del forms_df
            gc.collect()

            result_df = self.merge_meal_preferences(result_df)
            result_df = self.add_qr_codes(result_df, qr_df, contact_ids)
            del qr_df
            gc.collect()

            result_df = self._enrich_households_if_needed(result_df)
            result_df = self._apply_attendee_sort(result_df)
            result_df = self._apply_sub_event_filters(result_df)

            logger.info("Preprocessing data values...")
            result_df = self.preprocessor.preprocess_dataframe(result_df)

            has_config = self.config is not None
            has_inclusion_list = (
                has_config
                and getattr(self.config, 'inclusion_list', None) is not None
            )
            has_date_filter = (
                has_config
                and getattr(self.config, 'created_on_datetime', None) is not None
            )

            if has_inclusion_list or has_date_filter:
                original_count = len(result_df)
                filter_conditions = []
                
                if has_inclusion_list:
                    logger.info(
                        "Adding Contact ID filter for %d specified IDs",
                        len(self.config.inclusion_list),
                    )
                    contact_id_condition = result_df['Contact ID'].isin(self.config.inclusion_list)
                    if 'Member ID' in result_df.columns:
                        member_id_condition = result_df['Member ID'].isin(self.config.inclusion_list)
                        contact_id_condition = contact_id_condition | member_id_condition
                        logger.info("Filtering by both Contact ID (GUID) and Member ID (ID-####) formats")
                    filter_conditions.append(contact_id_condition)
                
                if has_date_filter:
                    logger.info(
                        "Adding date filter for registrations on or after: %s",
                        self.config.created_on_datetime,
                    )
                    date_contact_ids = self._date_filter_contact_ids()
                    if date_contact_ids is not None:
                        filter_conditions.append(result_df['Contact ID'].isin(date_contact_ids))
                
                if filter_conditions:
                    if len(filter_conditions) == 1:
                        combined_filter = filter_conditions[0]
                    else:
                        # Use OR logic to combine conditions
                        combined_filter = filter_conditions[0]
                        for condition in filter_conditions[1:]:
                            combined_filter = combined_filter | condition
                    
                    result_df = result_df[combined_filter]
                    logger.info(f"Applied filters: {original_count} -> {len(result_df)} contacts")
                    
                    # Log details for Contact ID filtering
                    if has_inclusion_list:
                        # Check matches in both Contact ID and Member ID columns
                        found_contact_ids = set(result_df['Contact ID'].unique()) & set(self.config.inclusion_list)
                        found_member_ids = set()
                        if 'Member ID' in result_df.columns:
                            found_member_ids = set(result_df['Member ID'].unique()) & set(self.config.inclusion_list)
                        
                        found_ids = found_contact_ids | found_member_ids
                        missing_ids = set(self.config.inclusion_list) - found_ids
                        
                        logger.info(f"ID filter matched {len(found_ids)} of {len(self.config.inclusion_list)} requested IDs")
                        if found_contact_ids:
                            logger.info(f"  - {len(found_contact_ids)} matched by Contact ID (GUID)")
                        if found_member_ids:
                            logger.info(f"  - {len(found_member_ids)} matched by Member ID (ID-####)")
                        
                        if missing_ids:
                            logger.warning(f"Could not find data for {len(missing_ids)} IDs:")
                            for missing_id in list(missing_ids)[:10]:  # Show first 10 to avoid spam
                                logger.warning(f"  - {missing_id}")
                            if len(missing_ids) > 10:
                                logger.warning(f"  ... and {len(missing_ids) - 10} more")
                else:
                    logger.info("No valid filter conditions to apply")
            
            # Collect and generate statistics
            self.stats_reporter.collect_statistics(result_df)
            self.stats_reporter.generate_report()
            
            return result_df
            
        except Exception as e:
            logger.error(f"Error during processing: {str(e)}")
            raise

    def save_output(self, df: pd.DataFrame) -> None:
        """Save the merged data to an Excel file with appropriate name."""
        output_filename = (self.config.get_output_filename() if self.config 
                         else f'MAIL_MERGE_v3_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        # Save the DataFrame to Excel
        df.to_excel(output_filename, index=False)
        
        # Add filters and freeze top row using openpyxl
        wb = load_workbook(output_filename)
        ws = wb.active

        # Freeze the top row
        ws.freeze_panes = 'A2'  # This freezes row 1 (header row)
        
        # Add filters to the top row
        ws.auto_filter.ref = ws.dimensions

        # Set column widths based on the longest header or value in each column
        lengths = (
            pd.concat([
                df.columns.to_series().map(len),                   # header lengths
                df.astype(str).applymap(len).max(axis=0)           # max per column
            ], axis=1)
            .max(axis=1)                                           # take the larger of the two
        )
        
        for idx, (col_name, max_len) in enumerate(lengths.iteritems(), start=1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = max_len + 2

        
        # Save the workbook with the new formatting
        wb.save(output_filename)
        
        logger.info(f"Merged data saved to: {output_filename} (with frozen header and filters)")

def main(sub_event: Optional[str] = None):
    try:
        # Initialize processor
        processor = EventRegistrationProcessorV3()
        
        # Find latest files and load registration data to get main event
        files = processor.find_latest_files()
        reg_df = pd.read_excel(files[FileTypes.REGISTRATION])
        
        # Get main event (Convention 2025 - San Francisco)
        main_event = reg_df[reg_df['Status Reason'] == 'Paid']['Event '].iloc[0]
        
        # Initialize configuration if sub_event is specified
        config = PreprocessingConfig(
            main_event=main_event,
            sub_event=sub_event
        ) if sub_event else None
        
        # Reinitialize processor with configuration
        processor = EventRegistrationProcessorV3(config)
        
        # Process and merge all data
        merged_df = processor.transform_and_merge()
        
        # Save the output
        processor.save_output(merged_df)
        
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        logger.error("\nPlease ensure all required files are present in the current directory with the correct naming format:")
        logger.error("  - Registration List: *Registration List*.xlsx")
        logger.error("  - Seating Chart: *Seating Chart*.xlsx")
        logger.error("  - QR Codes: *QR Codes*.xlsx")
        logger.error("  - Form Responses: *(Form|From) Responses*.xlsx")

if __name__ == "__main__":
    main()